# This is a sample Python script.

# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.


import os;
from datetime import datetime;

import numpy as np;
import pandas as pd;
from fpdf import  FPDF;
import openpyxl;

directory = "C:\cataeds"



def returnCodUnic(codUnic):
    startLetters = codUnic[0:2];
    startIndex = len(codUnic) - 8;
    stopIndex = len(codUnic);
    lastDigits = codUnic[startIndex:stopIndex];
    codUnic = startLetters + lastDigits;
    return codUnic;


def returnPunctIncarcare(judetPunctIncarcare, localitatePunctIncarcare):
    punctIncarcare = judetPunctIncarcare + "," + localitatePunctIncarcare;
    return punctIncarcare


def returnNrInmatriculare(numarCapTractor, numarRemorca):
    if (numarRemorca == "-"):
        return numarCapTractor;
    else:
        return numarCapTractor + "," + numarRemorca;


def returnDataFrame(data):
    data.iterrows()
    newData = data[['Sortiment', 'Specie', 'Volum(mc)']];
    groupedData = newData.groupby(['Sortiment', 'Specie'], as_index=False).aggregate(np.sum);
    nameDict={"Volum(mc)":"Volumn"};
    groupedData=groupedData.rename(columns=nameDict);
    return groupedData;

def returnDataFrameDC(data):
    data.iterrows()
    newData = data[['Subsortiment', 'Specie', 'Volum(mc)']];
    groupedData = newData.groupby(['Subsortiment', 'Specie'], as_index=False).aggregate(np.sum);
    nameDict={"Volum(mc)":"Volumn"};
    groupedData=groupedData.rename(columns=nameDict);
    return groupedData;



def writeNewCsv(codUnic, provenienta, emitent, dataSiOra, punctIncarcare, nrInmatriculare, df, societateaEmitenta,
                gestiunea, societateaClient):
    wb=openpyxl.load_workbook('C:\cataeds\OutputExcel.xlsx');

    print("sadsdad")
    sheet=wb["NIR-Global-Val-fara-TVA"]
    sheet.cell(row=3, column=4).value = societateaClient;
    sheet.cell(row=4, column=4).value=gestiunea;

    sheet.cell(row=4,column=10).value=dataSiOra;
    sheet.cell(row=6, column=3).value="Subsemnatii, membrii ai comisiei de receptie, am receptionat valorile materiale furnizate de: " + societateaEmitenta+" , delegat: ………………............,auto nr.: "+nrInmatriculare+ " cu punctul de incarcare: "+punctIncarcare+", pe baza documentelor insotitoare: …Factura/Aviz/etc "+ codUnic +", constatand:";
    sheet.cell(row=11,column=3).value="lemn";
    sheet.cell(row=11, column=6).value = "M.C";
    sheet.cell(row=11,column=7).value=2.35;

    print(df);
    index=0;
    rowNumber=11;

    for row in df.iterrows():
        if codUnic.startswith('DC'):
         sheet.cell(row=rowNumber, column=3).value = df.Subsortiment[index]+" "+df.Specie[index];
        else:
         sheet.cell(row=rowNumber, column=3).value = df.Sortiment[index] + " " + df.Specie[index];

        sheet.cell(row=rowNumber, column=6).value = "M.C";
        sheet.cell(row=rowNumber, column=7).value = df.Volumn[index];
        index=index+1;
        rowNumber=rowNumber+1;


    print("\\");
    filename="C:\outputs\\"+codUnic+".xlsx";
    print(filename);
    wb.save(filename);





for filename in os.scandir(directory):
    if filename.is_file():
        data = pd.read_csv(filename);
        codUnic = data['Cod unic'].values[0];
        provenienta = data['Provenienta'].values[0];
        emitent = data['Emitent - nume'].values[0];
        dataSiOra = data['Data si ora emiterii codului unic'].values[0];
        judetPunctIncarcare = data['Punct incarcare - Judet'].values[0];
        localitatePunctIncarcare = data['Punct incarcare - Localitate'].values[0];
        numarCapTractor = data['Cap tractor'].values[0];
        numarRemorca = data['Remorca'].values[0];
        gestiunea = data['Punct descarcare - Depozit'].values[0];
        societateaClient = data['Destinatar - Nume'].values[0];
        societateaEmitenta = data['Emitent - nume'].values[0];
        dataSiOra = datetime.strptime(dataSiOra, '%d-%m-%Y %H:%M:%S').date().strftime('%d-%m-%Y')
        codUnic = returnCodUnic(codUnic);
        punctIncarcare = returnPunctIncarcare(judetPunctIncarcare, localitatePunctIncarcare);
        nrInmatriculare = returnNrInmatriculare(numarCapTractor, numarRemorca);
        if codUnic.startswith('DC'):
         df = returnDataFrameDC(data);
        else:
            df=returnDataFrame(data);
        writeNewCsv(codUnic, provenienta, emitent, dataSiOra, punctIncarcare, nrInmatriculare, df, societateaEmitenta,
                    gestiunea, societateaClient);
